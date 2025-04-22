from models.loader import load_model_and_processor
from models.utils import get_visual_type, VALID_DATA_FORMAT_STRING
import os
from tqdm import tqdm
import json
import torch
from transformers import AutoTokenizer, AutoModelForCausalLM, GenerationConfig
import pandas as pd
import re
import gc
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
torch.backends.cudnn.enabled = False

# 加载xls文件，包含图片列
def process_input(file_path, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 判断工作表是否存在，不存在则使用第一个工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    
    # 获取表头所在行（假设第一行为表头），记录各列标题对应的列号
    header = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value is not None:
            header[cell_value] = col

    # 检查必须的列是否存在
    required_columns = ['缩略图', '抓拍时间', '抓拍地点']
    for col_name in required_columns:
        if col_name not in header:
            raise ValueError(f"找不到必需的列: {col_name}")
    
    # 获取各列的列号
    img_col = header['缩略图']
    time_col = header['抓拍时间']
    trace_col = header['抓拍地点']
    
    # 创建图片加载器，用于提取单元格中的图片
    image_loader = SheetImageLoader(ws)
    image_column = []
    time_column = []
    trace_column = []
    
    # 从第二行开始（跳过表头），并按照逆序处理（即列表中第一项为最后一行数据）
    for row in range(ws.max_row, 1, -1):
        # 生成“缩略图”所在单元格的坐标
        cell_coord = get_column_letter(img_col) + str(row)
        if image_loader.image_in(cell_coord):
            img = image_loader.get(cell_coord)
        else:
            img = None
        image_column.append(img)
        
        # 读取“抓拍时间”与“抓拍地点”对应单元格的值
        time_val = ws.cell(row=row, column=time_col).value
        trace_val = ws.cell(row=row, column=trace_col).value
        time_column.append(time_val)
        trace_column.append(trace_val)
    
    return image_column[:-11:-1], time_column[:-11:-1], trace_column[:-11:-1]

# vlm处理一个图片，输出对应的描述
def process_vlm(model, processor, prompt, video_file, generate_kwargs):
    inputs = processor(prompt, images=[video_file], edit_prompt=True, return_prompt=True)
    if 'prompt' in inputs:
        inputs.pop('prompt')
    inputs = {k:v.to(model.device) for k,v in inputs.items() if v is not None}
    outputs = model.generate(
        **inputs,
        **generate_kwargs,
    )
    output_text = processor.tokenizer.decode(outputs[0][inputs['input_ids'][0].shape[0]:], skip_special_tokens=True)
    return output_text

# llm的调用
def process_llm(llm, llm_tokenizer, prompt):
    messages = [{"role": "user", "content": prompt}]
    text = llm_tokenizer.apply_chat_template(
        messages,
        tokenize=False,
        add_generation_prompt=True
    )
    llm_inputs = llm_tokenizer([text], return_tensors="pt").to(llm.device)
    generated_ids = llm.generate(
        llm_inputs.input_ids,
        max_new_tokens=512
    )
    generated_ids = [
        output_ids[len(input_ids):] for input_ids, output_ids in zip(llm_inputs.input_ids, generated_ids)
    ]
    output_text = llm_tokenizer.batch_decode(generated_ids, skip_special_tokens=True)[0]
    return output_text

# 组织llm的prompt
def generate_prompt(images, times, traces):

    # 检查三个列表的长度是否一致
    if not (len(images) == len(times) == len(traces)):
        raise ValueError("prompt的输入中，三个列表的长度必须一致！")

    # 构造 prompt 的开头部分，说明任务要求
    prompt_lines = [
        "下面会提供针对一个人物A拍摄到的轨迹，请你用A代之这个人，要求进行详细分析和总结，并输出以下内容：",
        "1. 整个轨迹中A的穿着和外貌；",
        "2. 按时间顺序描述A的动作；",
        "3. 按时间顺序描述A的移动轨迹；",
        "4. 推断A在这个轨迹中的行为意图；",
        "5. 总结A是否存在危险行为，或含有危险信息。"
        "",
        "下面是这个轨迹的各个节点的信息："
    ]

    # 按编号组织每张图片的信息
    for idx, (img, t, loc) in enumerate(zip(images, times, traces), start=1):
        prompt_lines.append(f"节点{idx}：")
        prompt_lines.append(f"{img}")
        prompt_lines.append(f"  时间：{t}")
        prompt_lines.append(f"  地点：{loc}")
        prompt_lines.append("")  # 添加空行分隔

    # 拼接所有行，形成最终 prompt
    prompt_lines.append("请基于上述信息进行详细分析和总结。")
    prompt = "\n".join(prompt_lines)
    return prompt


def run():
    # 输入参数
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--model_name_or_path', type=str, help='Path to vlm, recommend PEAR')
    parser.add_argument('--llm_path', type=str, help='Path to llm, recommand Qwen2.5')
    parser.add_argument('--input_path', type=str, help='Path to trace file')
    parser.add_argument('--sheet_name', type=str, default="Sheet1", help='Excel file sheet name')
    parser.add_argument('--output_path', type=str, default="./debug.json", help='Path to save result, has to be a json file')
    parser.add_argument("--max_samples", type=int, default=0, help="Limit sample num.")
    parser.add_argument("--max_n_frames", type=int, default=8, help="Max number of frames to apply average sampling from the given video.")
    parser.add_argument("--max_new_tokens", type=int, default=512, help="max number of generated tokens")
    parser.add_argument("--top_p", type=float, default=1, help="Top_p sampling")
    parser.add_argument("--temperature", type=float, default=0, help="Set temperature > 0 to enable sampling generation.")
    args = parser.parse_args()

    # 加载vlm
    model, processor = load_model_and_processor(args.model_name_or_path, max_n_frames=args.max_n_frames, attn_implementation="flash_attention_2")
    generate_kwargs = {
        "do_sample": True if args.temperature > 0 else False,
        "max_new_tokens": args.max_new_tokens,
        "top_p": args.top_p,
        "temperature": args.temperature,
        "use_cache": True
    }

    # 检测输入与输出路径规范
    assert os.path.exists(args.input_path), f"input_path not exist: {args.input_path}"
    assert args.input_path.endswith('.xlsx'), f"invalid excel file name: {args.input_path}"
    
    assert os.path.exists(os.path.dirname(args.output_path)), f"output_path not exist: {args.output_path}"
    assert args.output_path.endswith('.json'), f"invalid json file name: {args.output_path}"

    # 处理输入
    images, times, traces = process_input(args.input_path, args.sheet_name)

    # 对每张图片生成画像
    portraits = []
    for image in tqdm(images, desc="生成画像中..."):
        prompt = " USER: Please describe:\n1. The person's appearance, wearing, and other important charateristic.\n2. The person's each action in order and the objects the person interacted with.\n3. The person's mood, feeling, emotion or intention if can be detected.</s> ASSISTANT: "
        prompt = "<image>\n" + prompt.replace("<image>", "").replace("<video>", "")
        pred = process_vlm(model, processor, prompt, image, generate_kwargs)
        portrait = ""
        if pred is not None:
            parts = re.split(r'\.(?!\s)', pred)
            portrait += '  外貌特点：'
            portrait += parts[0]
            portrait += '\n  行为动作：'
            portrait += parts[1]
            portrait += '\n  情感意图：'
            portrait += parts[2]
        portraits.append(portrait)

    # 释放显存
    model = model.cpu()
    del model
    gc.collect()
    torch.cuda.empty_cache()

    # 加载llm
    llm_name = args.llm_path
    llm_tokenizer = AutoTokenizer.from_pretrained(llm_name)
    llm = AutoModelForCausalLM.from_pretrained(llm_name, torch_dtype="auto", device_map="auto")

    # 分析总结
    prompt = generate_prompt(portraits, times, traces)
    analysis = process_llm(llm, llm_tokenizer, prompt)
    with open(args.output_path, "w", encoding="utf-8") as json_file:
        json.dump({'提问':prompt, '总结':analysis}, json_file, ensure_ascii=False, indent=4)

if __name__ == "__main__":
    run()